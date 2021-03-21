from openpyxl import load_workbook
# Giving the path of the excel file
workbook = load_workbook(r"C:\Users\KIIT\Desktop\Python_Practice\Train_Data_New.xlsx")
# Getting the sheet names
sheet_list=workbook.get_sheet_names()
# Printing the total sheets
print(sheet_list)   

# Defining function cell_v. In this cell value will be passed.
def cell_v(sheetname,row_value,column_value):
    #print(len(a))
    return (sheetname.cell(row=row_value, column=column_value).value)

# Defining function passenger ID and passing three arguments
def pass_id(workbook, sheet_name, passenger_id):
    mastersheet = workbook["Mastersheet"]
    for i in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=i, column=1).value == passenger_id):
            # Appending mastersheets
            mastersheet.append([cell_v(sheet_name,1,1), cell_v(sheet_name,i,1)])
            for j in range(2, sheet_name.max_column + 1):
                mastersheet.append([cell_v(sheet_name,1,j), cell_v(sheet_name,i,j)])
                workbook.save(r"C:\Users\KIIT\Desktop\Python_Practice\Train_Data_New.xlsx")


# workbook = load_workbook(r"D:\Python_Practice-1\18-03-2021\Train_Data.xlsx")
# # Getting the sheet names
# sheet_list=workbook.get_sheet_names()
# # Printing the total sheets
# print(sheet_list)    
# for 1

# Defining function passenger name and passing three arguments
def pass_name(workbook, sheet_name, passenger_name):
    mastersheet = workbook["Mastersheet"]
    for i in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=i, column=1).value == passenger_name):
            # Appending mastersheets
            mastersheet.append([cell_v(sheet_name,1,1), cell_v(sheet_name,i,1)])
            for j in range(2, sheet_name.max_column + 1):
                mastersheet.append([cell_v(sheet_name,1,j), cell_v(sheet_name,i,j)])
                workbook.save(r"C:\Users\KIIT\Desktop\Python_Practice\Train_Data_New.xlsx")


# Creating Mastersheet in excel file
if 'Mastersheet' in workbook.sheetnames:
    mas = workbook['Mastersheet']
# If mastersheet is there, it will remove.    
    workbook.remove(mas)   
 # It will again create a new mastersheet.                            
workbook.create_sheet("Mastersheet")              

# User will input the passenger ID
passenger_id = int(input("Enter the Passenger ID : ")) 
# User will input the passenger name
passenger_name = input("Enter the Passenger name : ")
# Getting the length.
for k in range(0, len(sheet_list)):
    sheetName = workbook[sheet_list[k]]
    # calling the passenger id function
    pass_id(workbook, sheetName, passenger_id)
    # calling the passenger name function
    pass_name(workbook, sheetName, passenger_name)

# for 2
# Creating another mastersheet with total number of rows and columns.
if 'Mastersheet1' in sheet_list:
    mas1 = workbook['Mastersheet1']
# If mastersheet is there, it will remove.    
    workbook.remove(mas1)   
# It will again create a new mastersheet.                                
workbook.create_sheet("Mastersheet1")

# Getting maximum row count
row_num=workbook['Mastersheet'].max_row
# Getting maximum column count
column2=workbook['Mastersheet'].max_column


mas1=workbook['Mastersheet1']
# Printing total number of rows in mastersheet 2
mas1.cell(row=1,column=1).value="Total number of rows"
# Printing total number of columns in mastersheet 2
mas1.cell(row=3,column=1).value="Total number of columns"
# Getting row value 
mas1.cell(row=1,column=2).value=row_num
# Getting column value
mas1.cell(row=3,column=2).value=column2
# Saving the file.
workbook.save(r"C:\Users\KIIT\Desktop\Python_Practice\Train_Data_New.xlsx")
# Printing total number of rows.
print("Total number of row is : ",row_num)
# Printing total number of column.
print("Total number of column is : ",column2)

# for k in range(0, len(sheet_list)):
#     row_count = sheetName.max_row
#     column_count = sheetName.max_column



