def get_data_PSNumber(workbook, sheet_name, user_input):
    mastersheet = workbook.get_sheet_by_name("mastersheet")
    for r in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=r, column=1).value == ps_number):
            mastersheet.append([cell_value(sheet_name,1,1), cell_value(sheet_name,r,1)])
            for c in range(2, sheet_name.max_column + 1):
                mastersheet.append([cell_value(sheet_name,1,c), cell_value(sheet_name,r,c)])
                workbook.save(r"D:\Python_Practice-1\18-03-2021\Demo.xlsx")

 


def get_data_Name(workbook, sheet_name, name):
    mastersheet = workbook.get_sheet_by_name("mastersheet")
    for r in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=r, column=2).value == name):
            mastersheet.append([cell_value(sheet_name,1,1), cell_value(sheet_name,r,1)])
            for c in range(2, sheet_name.max_column + 1):
                mastersheet.append([cell_value(sheet_name,1,c), cell_value(sheet_name,r,c)])
                workbook.save(r"D:\Python_Practice-1\18-03-2021\Demo.xlsx")

 

def get_data_Email(workbook, sheet_name, email):
    mastersheet = workbook.get_sheet_by_name("mastersheet")
    for r in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=r, column=3).value == email):
            mastersheet.append([cell_value(sheet_name,1,1), cell_value(sheet_name,r,1)])
            for c in range(2, sheet_name.max_column + 1):
                mastersheet.append([cell_value(sheet_name,1,c), cell_value(sheet_name,r,c)])
                workbook.save(r"D:\Python_Practice-1\18-03-2021\Demo.xlsx")

 

def cell_value(wo,ro, co):
    return (wo.cell(row=ro, column=co).value)

 

from openpyxl import load_workbook
workbook = load_workbook(r"D:\Python_Practice-1\18-03-2021\Demo.xlsx")
sheet_list=workbook.get_sheet_names()

 

if 'mastersheet' in workbook.sheetnames:
    msheet=workbook.get_sheet_by_name('mastersheet')
    jkl = workbook['mastersheet']
    workbook.remove(jkl)
workbook.create_sheet("mastersheet")

 

# user_input = int(input("Give a PS Number"))
#if(user_input==1):
ps_number = int(input("Enter PS number : "))
for i in range(0, len(sheet_list)):
    sheetName = workbook.get_sheet_by_name(sheet_list[i])
    get_data_PSNumber(workbook, sheetName, ps_number)
# elif (user_input==2):
#         name = input("enter PS name \n")
#         for i in range(0, len(sheet_list)):
#             sheetName = workbook.get_sheet_by_name(sheet_list[i])
#             get_data_Name(workbook, sheetName, name)

 

# elif (user_input==3):
#         email = input("enter ps Email Id \n")
#         for i in range(0, len(sheet_list)):
#             sheetName = workbook.get_sheet_by_name(sheet_list[i])
#             get_data_Email(workbook, sheetName, email)

 

# else:
#     print("Enter Please Input Valid\n")
        
