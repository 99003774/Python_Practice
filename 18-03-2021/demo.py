from openpyxl import load_workbook
# Giving the path of the excel file
workbook = load_workbook(r"D:\Python_Practice-1\18-03-2021\Train_Data_New.xlsx")
# Getting the sheet names
sheet_list=workbook.get_sheet_names()
# Printing the total sheets
print(sheet_list)   

# Defining function cell_v. In this cell value will be passed.
def cell_v(sheetname,row_value,column_value):
    #print(len(a))
    return (sheetname.cell(row=row_value, column=column_value).value)

# Defining function train number and passing three arguments
def train_number(workbook, sheet_name, train_num):
    mastersheet = workbook["Mastersheet"]
    for i in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=i, column=1).value == train_num):
            # Appending mastersheets
            mastersheet.append([cell_v(sheet_name,1,1), cell_v(sheet_name,i,1)])
            for j in range(2, sheet_name.max_column + 1):
                mastersheet.append([cell_v(sheet_name,1,j), cell_v(sheet_name,i,j)])
                workbook.save(r"D:\Python_Practice-1\18-03-2021\Train_Data_New.xlsx")
# workbook = load_workbook(r"D:\Python_Practice-1\18-03-2021\Train_Data.xlsx")
# # Getting the sheet names
# sheet_list=workbook.get_sheet_names()
# # Printing the total sheets
# print(sheet_list)    
# for 1

# Creating Mastersheet in excel file
if 'Mastersheet' in workbook.sheetnames:
    mas = workbook['Mastersheet']
# If mastersheet is there, it will remove.    
    workbook.remove(mas)   
 # It will again create a new mastersheet.                            
workbook.create_sheet("Mastersheet")              

# User will input the train number
train_num = int(input("Enter the train number : ")) 
# User will input the passenger name
pass_name = input("Enter the passenger name : ")
# Getting the length.
for k in range(0, len(sheet_list)):
    sheetName = workbook[sheet_list[k]]
    train_number(workbook, sheetName, train_num)


# for 2
# Creating another mastersheet with total number of rows and columns.
if 'Mastersheet1' in sheet_list:
    mas1 = workbook['Mastersheet1']
# If mastersheet is there, it will remove.    
    workbook.remove(mas1)   
# It will again create a new mastersheet.                                
workbook.create_sheet("Mastersheet1")

# Getting maximum row count
row_num=workbook['Mastersheet'].max_row
# Getting maximum column count
column2=workbook['Mastersheet'].max_column


mas1=workbook['Mastersheet1']
# Printing total number of rows in mastersheet 2
mas1.cell(row=1,column=1).value="Total number of rows"
# Printing total number of columns in mastersheet 2
mas1.cell(row=3,column=1).value="Total number of columns"
# Getting row value 
mas1.cell(row=1,column=2).value=row_num
# Getting column value
mas1.cell(row=3,column=2).value=column2
# Saving the file.
workbook.save(r"D:\Python_Practice-1\18-03-2021\Train_Data_New.xlsx")
# Printing total number of rows.
print("Total number of row is : ",row_num)
# Printing total number of column.
print("Total number of column is : ",column2)

# for k in range(0, len(sheet_list)):
#     row_count = sheetName.max_row
#     column_count = sheetName.max_column




# if 'Mastersheet2' in workbook.sheetnames:




## iteration to find the last row with values in it
# nrows = ws.max_row
# if nrows > 1000:
#     nrows = 1000

# lastrow = 0
# while True:
#     if ws.cell(nrows, 3).value != None:
#         lastrow = nrows
#         break
#     else:
#         nrows -= 1

# print(nrows)
# print(lastrow)


# number_of_rows = sheet_obj.max_row
#     last_row_index_with_data = 0
    
#     while True:
#         if sheet_obj.cell(number_of_rows, 3).value != None:
#             last_row_index_with_data = number_of_rows
#             break
#         else:
#             number_of_rows -= 1

